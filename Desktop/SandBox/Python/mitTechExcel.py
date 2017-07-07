import openpyxl, os, re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from copy import copy

#os.getcwd() will get working directory		os.chdir() will change working directory

## ---------------- Key ---------------- ##
#file_name = "Mitec example style BP"
sheet_name = "Budget & Forecast"
file_name = "Planning 2017-2021_Adjusted_v06"
#sheet_name = "Sheet1"



match_maker = {1 : re.compile('Direct Labor', re.IGNORECASE), 2 : re.compile('Variable costs', re.IGNORECASE), 3 : re.compile('Material Cost', re.IGNORECASE), 4 : re.compile('Staff Cost [SG&A]*', re.IGNORECASE), 5 : re.compile('Maintenance costs', re.IGNORECASE), 6 : re.compile('Rent and rates', re.IGNORECASE), 7 : re.compile('Overhead costs', re.IGNORECASE)}

things_to_insert = {1 : 'Direct Labor', 2 : 'Variable costs', 3 : 'Material Cost', 4 : 'Staff Cost', 5 : 'Maintenance', 6 : 'Rent', 7 : 'Overhead'}

#------------------------------------------------------------------------------ #

##------------------------------- Copy cell ---------------------------------- ##
def copyCell(from_row, to_row, from_col, to_col, new_sheet, old_sheet):
	oc = old_sheet.cell(row = from_row, column = from_col)
	nc = new_sheet.cell(row = to_row, column = to_col)
	# Copy Value
	old_value = old_sheet.cell(row = from_row, column = from_col).value
	new_sheet.cell(row = to_row, column = to_col).value = old_value

	## Copy Format 
	# Font
	font = copy(oc.font)
	nc.font = font
	# Fill
	fill = copy(oc.fill)
	nc.fill = fill
	# Border
	border = copy(oc.border)
	nc.border = border
	# Alignment
	alignment = copy(oc.alignment)
	nc.alignment = alignment
	# number_format
	number_format = copy(oc.number_format)
	nc.number_format = number_format

	
	return
#------------------------------------------------------------------------------ #

##------------------------ Inserting row function --------------------------- ##
def insertRows():
	
	my_start = row_to_insert_at
	my_rows = num_rows_to_insert
		
	print('Inserting rows... ')
	
	old_sheet = wb.get_sheet_by_name(sheet_name)
	mcol = old_sheet.max_column
	mrow = old_sheet.max_row
	old_sheet.title = (sheet_name + '.5')
	wb.create_sheet(index=0, title=sheet_name)
	
	new_sheet = wb.get_sheet_by_name(sheet_name)
	
	for row_num in range(1, my_start): 
		for col_num in range(1, mcol + 1):
			copyCell(row_num, row_num, col_num, col_num, new_sheet, old_sheet)
		print("Current row is: " + str(row_num))
	
	print("Next for loop coming...")   
	for row_num in range(my_start + my_rows, mrow + my_rows):
		for col_num in range(1, mcol + 1):
			copyCell(row_num - my_rows, row_num, col_num, col_num, new_sheet, old_sheet)
		
		print("Current row is: " + str(row_num) + " old cell value is: " + str(old_sheet.cell(row = row_num, column = col_num).value))
	return
# ------------------------------------------------------------------------------ #

#------------------- Add the values to the rows inserted ----------------------- #
#checker = "="
def addNamesToRows(start_row, end_row, count_by):
	sheet = wb.get_sheet_by_name(sheet_name)
	count = 1
	for x in range(start_row, end_row, count_by):
		sheet.cell(row = x, column = 1).value = things_to_insert[count]
		print("Inserting : " + str(sheet.cell(row = x, column = 1).value) + " at row: " + str(x))
		if(count < 7):
			count += 1	
		else:
			break	

#------------------------------------------------------------------------------ #	

#----------------------------- Save File -------------------------------------- #
def saveMe(workbook):
	os.chdir("/Users/BlackHawk/Desktop/")
	print('Saving to directory ' + os.getcwd() + '\nSaving... ')
	workbook.save("temp_excel1.xlsx")
	return
#------------------------------------------------------------------------------ #

#----------------------------- Alan's Function ---------------------------------#
def MAI(file_name, file_sheet_name, directory_address, cell_range):
	wb_2 = openpyxl.load_workbook(directory_address + file_name + ".xlsx")
	worksheet = wb_2.get_sheet_by_name(file_sheet_name)
	
	names = []
	years = []
	dates = {}
	price_count_avg = {}



# initiate dictionaries
	for year in range(2017, 2021 + 1):
		dates[year] = {}
		price_count_avg[year] = {}
	
	for x in cell_range:
		co_count = from_col_other_sheet
		cell_name = planning_sheet.cell(row = x, column = 1).value
		for year in range(2017, 2021 + 1):
			if(year in years and cell_name in dates[year]):
				print("ADDED: year = " + str(year) + " name: " + cell_name)
				dates[year][cell_name]['volume'] += planning_sheet.cell(row = x, column = co_count).value
				print("Volume is: " + str(dates[year][cell_name]['volume'])+'\n')
				dates[year][cell_name]['price'] +=  planning_sheet.cell(row = x, column = co_count+1).value	
				dates[year][cell_name]['total'] +=  planning_sheet.cell(row = x, column = co_count+2).value
				price_count_avg[year][cell_name] += 1
				
			else:
				dates[year][cell_name] = {}
				print("Created: year = " + str(year) + " name: " + cell_name)
				dates[year][cell_name]['volume'] = planning_sheet.cell(row = x, column = co_count).value
				print("Volume is: " + str(dates[year][cell_name]['volume']))
				dates[year][cell_name]['price'] =  planning_sheet.cell(row = x, column = co_count+1).value	
				print("price is: " + str(dates[year][cell_name]['price']))
				dates[year][cell_name]['total'] =  planning_sheet.cell(row = x, column = co_count+2).value
				print("total is: " + str(dates[year][cell_name]['total'])+'\n')
				if(cell_name not in names):
					names.append(cell_name)
				if(year not in years):
					years.append(year)
				price_count_avg[year][cell_name] = 1
	
			co_count += count_col_other_sheet
	
	#calculate price averages
	for year in range (2017, 2022):
		for co_name in names:
			num = dates[year][co_name]['price'] 
			den = price_count_avg[year][co_name]
			print("num: " + str(num) + " den: " + str(den))
			if(num):
				dates[year][co_name]['price'] = num / den
	

	##################   put the numbers in the table ----------------------
	table_col_begin = 12
	table_col_end = 15

	for x in range(table_col_begin, table_col_end + 1):
	#name
		worksheet.cell(row = x, column = 1).value = names[x - table_col_begin]
		worksheet.cell(row = x + lines_from_in_table_1, column = 1).value = names[x - table_col_begin]
		worksheet.cell(row = x + lines_from_in_table_2, column = 1).value = names[x - table_col_begin]
	
	#volume
		for year in range(3, 8):
			worksheet.cell(row = x, column = year).value = dates[2014 + year][worksheet.cell(row = x, column = 1).value]['volume']
	#price
			worksheet.cell(row = x + lines_from_in_table_1, column = year).value = dates[2014 + year][worksheet.cell(row = x + lines_from_in_table_1, column = 1).value]['price']
	
	#revenue
			worksheet.cell(row = x + lines_from_in_table_2, column = year).value = dates[2014 + year][worksheet.cell(row = x + lines_from_in_table_2, column = 1).value]['total']

	saveMe(wb_2)
	
	return	
	


###############################------ MAIN -----#################################
#num_rows_to_insert -= 1
wb = openpyxl.load_workbook('/Users/BlackHawk/Desktop/Jebsen Internship/Mitech/' + file_name + ".xlsx") #returns the value of Workbook data type
planning_sheet = wb.get_sheet_by_name(sheet_name)

#row_to_insert_at = 8
#num_rows_to_insert = 18
#insertRows()
#addNamesToRows()

#saveMe()

MAI_file = "MPI - REV BUD 17-19 (Jun17) v12"
MAI_dir = '/Users/BlackHawk/Desktop/Jebsen Internship/Mitech/'
MAI_sheet = "P&L"

print('Reading rows... ')
from_col_other_sheet = 11
count_col_other_sheet = 3

lines_from_in_table_1 = 7	
lines_from_in_table_2 = 13

MAI(MAI_file, MAI_sheet, MAI_dir)
MAG()



#------------ Alan Wants ---------------#
# Considate excel, add variables to asusmptions page with projects overhead
# For booked business, take planning file and input the volume, price, and calculate revenue which goes into P/L with material purchase price and material costs
